#!/usr/bin/env python3
"""Export a public iCalendar feed to an .xlsx sheet for one academic year."""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

import recurring_ical_events
import requests
from icalendar import Calendar
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


CHICAGO = ZoneInfo("America/Chicago")

# AY Y runs from summer Y through spring Y+1. Each entry: (name, (month, day), year_offset).
SEMESTERS = [
    ("SUMMER", (5, 16), 0),
    ("FALL",   (8, 21), 0),
    ("WINTER", (12, 15), 0),
    ("SPRING", (1, 10), 1),
]

DAY_NAMES = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]

COLUMNS = [
    "Summary", "Semester", "Day", "Week",
    "Date", "Start Time", "End Time", "All-Day",
    "Location", "Description", "Status",
    "Recurring?", "Meet Link", "UID",
]
COLUMN_WIDTHS = [40, 8, 5, 6, 12, 10, 10, 8, 25, 40, 10, 10, 30, 32]


@dataclass
class EventRow:
    summary: str
    semester: str
    day: str
    week: int
    date: str
    start_time: str
    end_time: str
    all_day: str
    location: str
    description: str
    status: str
    recurring: str
    meet_link: str
    uid: str


def current_ay_start_year(today: date | None = None) -> int:
    today = today or date.today()
    # AY starts May 16. Anything before that is still the prior AY.
    return today.year if (today.month, today.day) >= (5, 16) else today.year - 1


def semester_start_date(name: str, ay: int) -> date:
    for sem_name, (month, day), offset in SEMESTERS:
        if sem_name == name:
            return date(ay + offset, month, day)
    raise ValueError(f"Unknown semester: {name}")


def ay_window(ay: int) -> tuple[date, date]:
    return date(ay, 5, 16), date(ay + 1, 5, 15)


def semester_of(d: date, ay: int) -> str | None:
    boundaries = [(name, semester_start_date(name, ay)) for name, _, _ in SEMESTERS]
    ay_end = ay_window(ay)[1]
    for i, (name, start) in enumerate(boundaries):
        end = boundaries[i + 1][1] - timedelta(days=1) if i + 1 < len(boundaries) else ay_end
        if start <= d <= end:
            return name
    return None


def week_number(d: date, semester_start: date) -> int:
    # Week 1 = the first Sunday STRICTLY after the semester cutoff.
    # Everything from the cutoff up to that Sunday-minus-1 is Week 0.
    days_to_sunday = (6 - semester_start.weekday()) % 7  # Python: Mon=0..Sun=6
    if days_to_sunday == 0:
        days_to_sunday = 7
    week1_start = semester_start + timedelta(days=days_to_sunday)
    if d < week1_start:
        return 0
    return ((d - week1_start).days // 7) + 1


def day_name(d: date) -> str:
    return DAY_NAMES[d.weekday()]


def to_local_date(dt: date | datetime) -> date:
    if isinstance(dt, datetime):
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=CHICAGO)
        return dt.astimezone(CHICAGO).date()
    return dt


def to_local_time_str(dt: date | datetime) -> str:
    if not isinstance(dt, datetime):
        return ""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=CHICAGO)
    return dt.astimezone(CHICAGO).strftime("%H:%M")


def is_all_day(dt: date | datetime) -> bool:
    return not isinstance(dt, datetime)


def prop_str(event, key: str) -> str:
    value = event.get(key)
    return str(value) if value is not None else ""


def fetch_calendar(url: str) -> Calendar:
    response = requests.get(url, timeout=30)
    response.raise_for_status()
    return Calendar.from_ical(response.content)


def recurring_uids(calendar: Calendar) -> set[str]:
    return {
        str(component["UID"])
        for component in calendar.walk("VEVENT")
        if component.get("RRULE") and component.get("UID")
    }


def build_rows(
    calendar: Calendar,
    ay: int,
    exclude_patterns: list[str] | None = None,
) -> tuple[list[EventRow], int]:
    start_date, end_date = ay_window(ay)
    start_dt = datetime.combine(start_date, datetime.min.time(), tzinfo=CHICAGO)
    end_dt = datetime.combine(end_date, datetime.max.time(), tzinfo=CHICAGO)

    rec_uids = recurring_uids(calendar)
    expanded = recurring_ical_events.of(calendar).between(start_dt, end_dt)

    patterns = [p.lower() for p in (exclude_patterns or [])]

    rows: list[EventRow] = []
    excluded = 0
    for event in expanded:
        dtstart_prop = event.get("DTSTART")
        if dtstart_prop is None:
            continue
        dtstart = dtstart_prop.dt
        dtend_prop = event.get("DTEND")
        dtend = dtend_prop.dt if dtend_prop is not None else dtstart

        local_date = to_local_date(dtstart)
        semester = semester_of(local_date, ay)
        if semester is None:
            continue

        summary = prop_str(event, "SUMMARY")
        if patterns and any(p in summary.lower() for p in patterns):
            excluded += 1
            continue

        all_day = is_all_day(dtstart)
        uid = prop_str(event, "UID")

        rows.append(EventRow(
            summary=summary,
            semester=semester,
            day=day_name(local_date),
            week=week_number(local_date, semester_start_date(semester, ay)),
            date=local_date.isoformat(),
            start_time="" if all_day else to_local_time_str(dtstart),
            end_time="" if all_day else to_local_time_str(dtend),
            all_day="Yes" if all_day else "No",
            location=prop_str(event, "LOCATION"),
            description=prop_str(event, "DESCRIPTION"),
            status=prop_str(event, "STATUS"),
            recurring="Yes" if uid in rec_uids else "No",
            meet_link=prop_str(event, "X-GOOGLE-CONFERENCE"),
            uid=uid,
        ))

    rows.sort(key=lambda r: (r.date, r.start_time))
    return rows, excluded


def write_xlsx(rows: list[EventRow], output_path: Path, ay: int) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"AY {ay}-{str(ay + 1)[-2:]}"

    sheet.append(COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"

    for row in rows:
        sheet.append([
            row.summary, row.semester, row.day, row.week,
            row.date, row.start_time, row.end_time, row.all_day,
            row.location, row.description, row.status,
            row.recurring, row.meet_link, row.uid,
        ])

    for i, width in enumerate(COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width

    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export a public .ics calendar to an .xlsx sheet for one academic year.",
    )
    parser.add_argument("url", help="Public .ics calendar URL")
    parser.add_argument(
        "--year", type=int, default=None,
        help="Starting year of the academic year (e.g., 2026 = AY 2026-27). "
             "Defaults to the current AY based on today's date.",
    )
    parser.add_argument(
        "-o", "--output", type=Path, default=None,
        help="Output .xlsx path. Defaults to calendar_AY<start>-<end>.xlsx",
    )
    parser.add_argument(
        "--exclude", action="append", default=[], metavar="PATTERN",
        help="Skip events whose summary contains PATTERN (case-insensitive). "
             "Repeatable, e.g. --exclude birthday --exclude 'happy hour'.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    ay = args.year if args.year is not None else current_ay_start_year()
    output = args.output or Path(f"calendar_AY{ay}-{str(ay + 1)[-2:]}.xlsx")

    print(f"Fetching {args.url}")
    calendar = fetch_calendar(args.url)

    print(f"Building rows for AY {ay}-{str(ay + 1)[-2:]}")
    rows, excluded = build_rows(calendar, ay, exclude_patterns=args.exclude)
    if args.exclude:
        print(f"  {len(rows)} events in window ({excluded} excluded by filter)")
    else:
        print(f"  {len(rows)} events in window")

    write_xlsx(rows, output, ay)
    print(f"Wrote {output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
